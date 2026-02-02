import io
import logging

from rest_framework import viewsets, status
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.utils import timezone
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from .models import Bot, SubscriptionPlan, User, Subscription, Payment, PaymentMethod, Messages, BotPlan
from .utils import send_test_message_sync, format_message, send_telegram_message_http
from .bot_api import add_subscription_to_bot, delete_subscription_from_bot
from .serializers import (
    BotSerializer,
    SubscriptionPlanSerializer,
    UserSerializer,
    SubscriptionCreateSerializer,
    SubscriptionReadSerializer,
    PaymentCreateSerializer,
    PaymentReadSerializer,
    PaymentMethodSerializer,
    MessagesSerializer,
    BotPlan, BotPlanSerializer
)


logger = logging.getLogger(__name__)


class FullPaymentReportView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        date_from_str = request.query_params.get("from")
        date_to_str = request.query_params.get("to")
        bot_id = request.query_params.get("bot_id")

        if date_from_str and date_to_str:
            date_from = timezone.make_aware(datetime.strptime(date_from_str, "%Y-%m-%d"))
            date_to = timezone.make_aware(datetime.strptime(date_to_str, "%Y-%m-%d"))
        else:
            now = timezone.now()
            date_from = timezone.make_aware(datetime(now.year, now.month, 1))
            date_to = now

        payments = Payment.objects.filter(created_at__range=[date_from, date_to], status="success")
        if bot_id:
            payments = payments.filter(id=bot_id)

        total_payments = payments.count()
        total_revenue = sum(p.amount for p in payments)

        by_method = {}
        by_bot = {}
        for p in payments:
            by_method.setdefault(p.method, {"count": 0, "amount": 0})
            by_method[p.method]["count"] += 1
            by_method[p.method]["amount"] += p.amount

            bot_name = p.bot.username if p.bot else "—"
            by_bot.setdefault(bot_name, {"count": 0, "amount": 0})
            by_bot[bot_name]["count"] += 1
            by_bot[bot_name]["amount"] += p.amount

        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Report"

        ws.append(["📊 Отчёт по оплатам"])
        ws.append(["Период", f"{date_from.date()} — {date_to.date()}"])
        ws.append(["Всего платежей", total_payments])
        ws.append(["Общая сумма", total_revenue])
        ws.append([])

        ws.append(["Разбивка по методам оплаты"])
        ws.append(["Метод", "Количество", "Сумма"])
        for m, v in by_method.items():
            ws.append([m, v["count"], v["amount"]])
        ws.append([])

        ws.append(["Разбивка по ботам"])
        ws.append(["Бот", "Количество", "Сумма"])
        for b, v in by_bot.items():
            ws.append([b, v["count"], v["amount"]])
        ws.append([])

        ws.append(["ID", "User (Telegram ID)", "Bot", "Plan", "Amount", "Method", "Status", "Дата покупки"])
        for p in payments:
            ws.append([
                p.id,
                p.user.telegram_id if p.user else "—",
                p.bot.username if p.bot else "—",
                p.subscription.plan.name if p.subscription and p.subscription.plan else "—",
                p.amount,
                p.method,
                p.status,
                p.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])

        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(e)
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"payments_report_{date_from.date()}_{date_to.date()}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class PaymentSummaryReportView(APIView):

    def get(self, request):
        date_from = request.query_params.get("from", "2025-10-01")
        date_to = request.query_params.get("to", "2025-10-30")

        payments = Payment.objects.filter(created_at__range=[date_from, date_to])

        total_revenue = payments.aggregate(Sum("amount"))["amount__sum"] or 0
        total_payments = payments.count()

        methods = payments.values("method").annotate(
            count=Count("id"),
            amount=Sum("amount")
        )

        bots = payments.values("bot__username").annotate(
            count=Count("id"),
            amount=Sum("amount")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments Summary"

        bold = Font(bold=True)

        ws.append(["Total Revenue", total_revenue])
        ws.append(["Total Payments", total_payments])
        ws.append([])
        ws.append(["Period", f"{date_from} - {date_to}"])
        ws.append([])
        ws.append([])

        ws.append(["By Method"])
        ws.append(["Method", "Count", "Amount"])
        for m in methods:
            ws.append([
                m["method"],
                m["count"],
                m["amount"],
            ])
        ws.append([])
        ws.append([])

        ws.append(["By Bot"])
        ws.append(["Bot", "Count", "Amount"])
        for b in bots:
            ws.append([
                b["bot__username"] or "-",
                b["count"],
                b["amount"],
            ])

        for cell in ws["A1":"B1"][0]:
            cell.font = bold

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = FileResponse(
            stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="payments_summary_{date_from}_{date_to}.xlsx"'
        return response



class BotViewSet(viewsets.ModelViewSet):
    queryset = Bot.objects.all()
    serializer_class = BotSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        username = self.request.query_params.get("username")
        if username:
            qs = qs.filter(username__iexact=username)
        return qs


class SubscriptionPlanViewSet(viewsets.ModelViewSet):
    queryset = SubscriptionPlan.objects.all()
    serializer_class = SubscriptionPlanSerializer

    def get_queryset(self):
        queryset = SubscriptionPlan.objects.all()
        bot_id = self.request.query_params.get("bot_id")
        if bot_id:
            queryset = queryset.filter(bot_id=bot_id)
        return queryset.order_by("duration_days")


class BotPlanViewSet(viewsets.ModelViewSet):
    """
    API для работы с планами ботов (связь Bot-Plan с ценами)

    GET /api/bot-plans/ - список всех планов ботов
    GET /api/bot-plans/{id}/ - получить план бота по ID
    GET /api/bot-plans/?bot_id=1 - планы для конкретного бота
    GET /api/bot-plans/?bot_id=1&is_active=true - только активные планы бота
    POST /api/bot-plans/ - привязать план к боту
    PUT /api/bot-plans/{id}/ - обновить цены/активность
    DELETE /api/bot-plans/{id}/ - удалить связь
    """
    queryset = BotPlan.objects.select_related('bot', 'plan').all()
    serializer_class = BotPlanSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        bot_id = self.request.query_params.get('bot_id')
        if bot_id:
            queryset = queryset.filter(bot_id=bot_id)

        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            is_active_bool = is_active.lower() in ('true', '1', 'yes')
            queryset = queryset.filter(is_active=is_active_bool)

        return queryset

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Массовое создание планов для бота

        POST /api/bot_plans/bulk_create/
        {
            "bot_id": 1,
            "plans": [
                {
                    "plan_id": 1,
                    "price_usdt": 10.00,
                    "price_stars": 100,
                    "is_active": true
                },
                {
                    "plan_id": 2,
                    "price_usdt": 25.00,
                    "price_stars": 250,
                    "is_active": true
                }
            ]
        }
        """
        bot_id = request.data.get('bot_id')
        plans_data = request.data.get('plans', [])

        if not bot_id:
            return Response(
                {"error": "bot_id обязателен"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            bot = Bot.objects.get(id=bot_id)
        except Bot.DoesNotExist:
            return Response(
                {"error": f"Бот с ID {bot_id} не найден"},
                status=status.HTTP_404_NOT_FOUND
            )

        created_plans = []
        errors = []

        for plan_data in plans_data:
            plan_id = plan_data.get('plan_id')

            try:
                plan = SubscriptionPlan.objects.get(id=plan_id)

                bot_plan, created = BotPlan.objects.update_or_create(
                    bot=bot,
                    plan=plan,
                    defaults={
                        'is_active': plan_data.get('is_active', True),
                    }
                )

                created_plans.append({
                    'id': bot_plan.id,
                    'plan': plan.name,
                    'created': created
                })

            except SubscriptionPlan.DoesNotExist:
                errors.append(f"План с ID {plan_id} не найден")
            except Exception as e:
                errors.append(f"Ошибка при создании плана {plan_id}: {str(e)}")

        return Response({
            'success': True,
            'created': len(created_plans),
            'plans': created_plans,
            'errors': errors
        })


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    lookup_field = "telegram_id"


class PaymentMethodViewSet(viewsets.ModelViewSet):
    queryset = PaymentMethod.objects.filter()
    serializer_class = PaymentMethodSerializer


class MessageViewSet(viewsets.ModelViewSet):
    queryset = Messages.objects.all()
    serializer_class = MessagesSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        identifier = self.request.query_params.get("identifier")

        if identifier:
            queryset = queryset.filter(identifier=identifier)

        return queryset


class SubscriptionViewSet(viewsets.ModelViewSet):
    queryset = Subscription.objects.all()

    def get_serializer_class(self):
        if self.action in ["list", "retrieve"]:
            return SubscriptionReadSerializer
        return SubscriptionCreateSerializer


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()

    def get_serializer_class(self):
        if self.action in ["create", "mock"]:
            return PaymentCreateSerializer
        return PaymentReadSerializer

    @action(detail=False, methods=["post"])
    def mock(self, request):
        telegram_id = request.data.get("telegram_id")
        bot_id = request.data.get("bot_id")
        plan_id = request.data.get("plan_id")
        method = request.data.get("method", "stub")

        user, _ = User.objects.get_or_create(telegram_id=telegram_id)
        bot = get_object_or_404(Bot, id=bot_id)
        plan = get_object_or_404(SubscriptionPlan, id=plan_id)

        start = timezone.now()
        end = start + timezone.timedelta(days=plan.duration_days) if plan.duration_days else None

        sub = Subscription.objects.create(
            user=user,
            bot=bot,
            plan=plan,
            start_date=start,
            end_date=end,
            is_active=True
        )

        payment = Payment.objects.create(
            user=user,
            bot=bot,
            subscription=sub,
            method=method,
            amount=plan.price_usdt,
            status="success"
        )

        return Response(
            PaymentReadSerializer(payment).data,
            status=201
        )


# --------- API for Bots ---------

@api_view(["GET"])
@permission_classes([AllowAny])
def is_subscribed(request):
    user_id = request.GET.get("user_id")
    bot_username = request.GET.get("botusername")

    bot = get_object_or_404(Bot, username=bot_username)
    user = User.objects.filter(telegram_id=user_id).first()

    if not user:
        return Response({"is_subscribed": False})

    sub = Subscription.objects.filter(user=user, bot=bot, is_active=True).order_by("-end_date").first()

    if not sub:
        return Response({"is_subscribed": False})

    return Response({
        "is_subscribed": True,
        "subscription_start_date": int(sub.start_date.timestamp()) if sub.start_date else None,
        "subscription_end_date": int(sub.end_date.timestamp()) if sub.end_date else None
    })


@api_view(["GET"])
def subscribers(request):
    bot_username = request.GET.get("botusername")
    bot = get_object_or_404(Bot, username=bot_username)

    subs = Subscription.objects.filter(bot=bot, is_active=True)
    user_ids = list(subs.values_list("user__telegram_id", flat=True))

    return Response(user_ids)


@api_view(['POST'])
def send_test_message(request):
    """
    API endpoint для отправки тестового сообщения пользователю

    POST /api/send-test-message/

    Body:
    {
        "user_id": 123456789,
        "message_text": "Текст сообщения"
    }

    Response:
    {
        "success": true,
        "message": "✅ Сообщение успешно отправлено пользователю 123456789"
    }
    """
    user_id = request.data.get('user_id')
    message_text = request.data.get('message_text')

    if not user_id:
        return Response(
            {
                "success": False,
                "message": "❌ Параметр 'user_id' обязателен"
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not message_text:
        return Response(
            {
                "success": False,
                "message": "❌ Параметр 'message_text' обязателен"
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user_id = int(user_id)
    except (ValueError, TypeError):
        return Response(
            {
                "success": False,
                "message": "❌ Параметр 'user_id' должен быть числом"
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    formatted_text = format_message(message_text)
    result = send_test_message_sync(user_id, formatted_text)

    if result['success']:
        return Response(result, status=status.HTTP_200_OK)
    else:
        return Response(result, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def send_notification(request):
    """
    API endpoint для отправки уведомления о покупке подписки пользователю

    POST /api/send-purchase-notification/

    Body:
    {
        "user_id": 123456789,
        "message_identifier": "subscription_purchased",
        "bot_id": 1,
        "plan_id": 5,
        "language": "ru"
    }

    Response:
    {
        "success": true,
        "message": "✅ Сообщение о покупке отправлено пользователю 123456789"
    }
    """
    user_id = request.data.get('user_id')
    bot_id = request.data.get('bot_id')
    plan_id = request.data.get('plan_id')
    message_identifier = request.data.get('message_identifier')
    language = request.data.get('language', 'ru')

    if not user_id:
        return Response(
            {"success": False, "message": "❌ Параметр 'user_id' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not bot_id:
        return Response(
            {"success": False, "message": "❌ Параметр 'bot_id' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not plan_id:
        return Response(
            {"success": False, "message": "❌ Параметр 'plan_id' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user_id = int(user_id)
    except (ValueError, TypeError):
        return Response(
            {"success": False, "message": "❌ Параметр 'user_id' должен быть числом"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        try:
            bot = Bot.objects.get(id=bot_id)
        except Bot.DoesNotExist:
            return Response(
                {"success": False, "message": f"❌ Бот с ID {bot_id} не найден"},
                status=status.HTTP_404_NOT_FOUND
            )

        if not bot.bot_token:
            return Response(
                {"success": False, "message": f"❌ У бота @{bot.username} не указан токен"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            plan = SubscriptionPlan.objects.get(id=plan_id)
        except SubscriptionPlan.DoesNotExist:
            return Response(
                {"success": False, "message": f"❌ План с ID {plan_id} не найден"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            message_template = Messages.objects.get(identifier=message_identifier)
        except Messages.DoesNotExist:
            return Response(
                {"success": False, "message": "❌ Сообщение 'subscription_purchased' не найдено в БД"},
                status=status.HTTP_404_NOT_FOUND
            )

        message_field = f'message_{language}'
        message_text = getattr(message_template, message_field, None)

        if not message_text:
            message_text = message_template.message_ru
            if not message_text:
                return Response(
                    {"success": False, "message": f"❌ Нет текста сообщения для языка {language}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        duration_days = plan.duration_days or 30
        end_date = datetime.now() + timedelta(days=duration_days)

        formatted_message = format_message(message_text, end_date)

        result = send_telegram_message_http(
            bot_token=bot.bot_token,
            chat_id=user_id,
            text=formatted_message
        )

        if result['success']:
            logger.info(f"✅ Сообщение о покупке отправлено пользователю {user_id} от бота @{bot.username}")
            return Response(result, status=status.HTTP_200_OK)
        else:
            logger.error(f"❌ Не удалось отправить сообщение пользователю {user_id}: {result['message']}")
            return Response(result, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    except Exception as e:
        logger.exception(f"Ошибка при отправке сообщения о покупке: {e}")
        return Response(
            {"success": False, "message": f"❌ Неожиданная ошибка: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
def sync_subscription_status(request):
    """
    API endpoint для синхронизации статуса подписки с ботом клиента

    POST /api/sync-subscription/

    Body:
    {
        "user_id": 123456789,
        "bot_id": 1,
        "action": "activate"  // или "deactivate"
    }

    Response:
    {
        "success": true,
        "message": "✅ Подписка активирована",
        "active": true
    }
    """
    user_id = request.data.get('user_id')
    bot_id = request.data.get('bot_id')
    event = request.data.get('action')  # "activate" | "deactivate"

    if not user_id:
        return Response(
            {"success": False, "message": "❌ Параметр 'user_id' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not bot_id:
        return Response(
            {"success": False, "message": "❌ Параметр 'bot_id' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not event:
        return Response(
            {"success": False, "message": "❌ Параметр 'action' обязателен"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if event not in ['activate', 'deactivate']:
        return Response(
            {"success": False, "message": "❌ Параметр 'action' должен быть 'activate' или 'deactivate'"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user_id = int(user_id)
    except (ValueError, TypeError):
        return Response(
            {"success": False, "message": "❌ Параметр 'user_id' должен быть числом"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        try:
            bot = Bot.objects.get(id=bot_id)
        except Bot.DoesNotExist:
            return Response(
                {"success": False, "message": f"❌ Бот с ID {bot_id} не найден"},
                status=status.HTTP_404_NOT_FOUND
            )

        if event == 'activate':
            result = add_subscription_to_bot(bot, user_id)
        else:  # deactivate
            result = delete_subscription_from_bot(bot, user_id)

        if result['success']:
            return Response(result, status=status.HTTP_200_OK)
        else:
            return Response(result, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    except Exception as e:
        logger.exception(f"Ошибка при синхронизации подписки: {e}")
        return Response(
            {"success": False, "message": f"❌ Неожиданная ошибка: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
